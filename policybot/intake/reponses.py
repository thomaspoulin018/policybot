"""Lecture d'un export Microsoft Forms (.xlsx) vers des `DemandeIAG`.

Ligne 1 : les intitulés de questions. Une ligne par réponse. L'appariement
d'une colonne se fait sur l'intitulé normalisé, jamais sur sa position :
Forms insère ses propres colonnes et l'ordre des questions peut bouger.

Une ligne illisible ne fait pas tomber le lot — elle est rejetée avec son
motif et les autres passent, comme une recherche Exa en échec n'arrête pas
les seize autres.
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re

from pydantic import BaseModel, Field, ValidationError

from policybot.intake.formulaire import (
    COLONNES_TECHNIQUES,
    CatalogueFormulaire,
    QuestionFormulaire,
    formulaire,
    normaliser_intitule,
)
from policybot.intake.schema import DemandeIAG


_VRAI = {"oui", "yes", "true", "vrai", "1", "x"}
_FAUX = {"non", "no", "false", "faux", "0", ""}
_SEPARATEUR_CHOIX = ";"
_FORMATS_DATE = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d")


class ReponseIllisibleError(ValueError):
    """Une cellule ne peut pas être traduite vers le champ qu'elle alimente."""


class RejetDemande(BaseModel):
    ligne: int
    motif: str


class LotReponses(BaseModel):
    """Le résultat d'une lecture : ce qui passe, ce qui est écarté, et pourquoi."""

    demandes: list[DemandeIAG] = Field(default_factory=list)
    rejets: list[RejetDemande] = Field(default_factory=list)
    colonnes_ignorees: list[str] = Field(default_factory=list)
    colonnes_manquantes: list[str] = Field(default_factory=list)
    lignes_lues: int = 0


def _texte(valeur: object) -> str:
    if valeur is None:
        return ""
    if isinstance(valeur, (datetime, date)):
        return valeur.isoformat()
    if isinstance(valeur, float) and valeur.is_integer():
        return str(int(valeur))
    return str(valeur).strip()


def _entier(valeur: object, question: QuestionFormulaire) -> int | None:
    if valeur is None or _texte(valeur) == "":
        return None
    if isinstance(valeur, bool):
        raise ReponseIllisibleError(
            f"question {question.numero} : un nombre était attendu."
        )
    if isinstance(valeur, (int, float)):
        return int(valeur)
    trouve = re.search(r"-?\d+", _texte(valeur))
    if trouve is None:
        raise ReponseIllisibleError(
            f"question {question.numero} : un nombre était attendu."
        )
    return int(trouve.group())


def _date(valeur: object, question: QuestionFormulaire) -> date | None:
    if valeur is None or _texte(valeur) == "":
        return None
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    brut = _texte(valeur)
    for format_ in _FORMATS_DATE:
        try:
            return datetime.strptime(brut, format_).date()
        except ValueError:
            continue
    raise ReponseIllisibleError(
        f"question {question.numero} : date illisible (formats acceptés : "
        + ", ".join(_FORMATS_DATE)
        + ")."
    )


def _booleen(valeur: object, question: QuestionFormulaire) -> bool:
    if isinstance(valeur, bool):
        return valeur
    brut = _texte(valeur).casefold()
    if brut in _VRAI:
        return True
    if brut in _FAUX:
        return False
    traduit = question.valeur_pour(brut)
    if traduit is not None and traduit.casefold() in _VRAI:
        return True
    if traduit is not None and traduit.casefold() in _FAUX:
        return False
    raise ReponseIllisibleError(
        f"question {question.numero} : réponse « {brut} » hors des choix oui / non."
    )


def _choix(valeur: object, question: QuestionFormulaire) -> str | None:
    brut = _texte(valeur)
    if not brut:
        return None
    traduit = question.valeur_pour(brut)
    if traduit is None:
        # Le libellé d'un choix est du texte de formulaire, pas une donnée du
        # demandeur : il peut être cité tel quel dans le motif de rejet.
        raise ReponseIllisibleError(
            f"question {question.numero} : réponse « {brut} » hors des choix proposés."
        )
    return traduit


def _choix_multiple(valeur: object, question: QuestionFormulaire) -> list[str]:
    brut = _texte(valeur)
    if not brut:
        return []
    retenus: list[str] = []
    for morceau in brut.split(_SEPARATEUR_CHOIX):
        libelle = morceau.strip()
        if not libelle:
            continue
        # Forms autorise une réponse « Autre » libre : on la conserve telle
        # quelle plutôt que de rejeter la demande.
        retenus.append(question.valeur_pour(libelle) or libelle)
    return retenus


def _valeur_pour(question: QuestionFormulaire, brut: object):
    if question.type in ("texte", "texte_long"):
        return _texte(brut)
    if question.type == "nombre":
        return _entier(brut, question)
    if question.type == "date":
        return _date(brut, question)
    if question.type == "oui_non":
        return _booleen(brut, question)
    if question.type == "choix_multiple":
        return _choix_multiple(brut, question)
    return _choix(brut, question)


def _lire_feuille(chemin: Path) -> list[list[object]]:
    from openpyxl import load_workbook

    classeur = load_workbook(filename=chemin, read_only=True, data_only=True)
    try:
        feuille = classeur[classeur.sheetnames[0]]
        return [list(ligne) for ligne in feuille.iter_rows(values_only=True)]
    finally:
        classeur.close()


def _apparier_colonnes(
    entetes: list[object], catalogue: CatalogueFormulaire
) -> tuple[dict[int, QuestionFormulaire], list[str], list[str]]:
    par_intitule = catalogue.par_intitule()
    techniques = {normaliser_intitule(nom) for nom in COLONNES_TECHNIQUES}
    colonnes: dict[int, QuestionFormulaire] = {}
    ignorees: list[str] = []
    for index, entete in enumerate(entetes):
        libelle = _texte(entete)
        if not libelle:
            continue
        cle = normaliser_intitule(libelle)
        question = par_intitule.get(cle)
        if question is None:
            if cle not in techniques:
                ignorees.append(libelle)
            continue
        colonnes[index] = question
    trouvees = {question.champ for question in colonnes.values()}
    manquantes = [q.intitule for q in catalogue.questions if q.champ not in trouvees]
    return colonnes, ignorees, manquantes


def lire_export(
    chemin: str | Path,
    catalogue: CatalogueFormulaire | None = None,
) -> LotReponses:
    """Lit un export Forms et rend les demandes valides et les rejets motivés."""
    catalogue = catalogue or formulaire()
    chemin = Path(chemin)
    if not chemin.is_file():
        raise FileNotFoundError(f"Export Microsoft Forms introuvable : {chemin}")

    lignes = _lire_feuille(chemin)
    if not lignes:
        return LotReponses()

    colonnes, ignorees, manquantes = _apparier_colonnes(lignes[0], catalogue)
    obligatoires_absentes = [
        question.intitule
        for question in catalogue.questions
        if question.obligatoire and question.champ not in {q.champ for q in colonnes.values()}
    ]

    lot = LotReponses(
        colonnes_ignorees=ignorees,
        colonnes_manquantes=manquantes,
        lignes_lues=0,
    )
    for numero_ligne, ligne in enumerate(lignes[1:], start=2):
        if all(_texte(cellule) == "" for cellule in ligne):
            continue
        lot.lignes_lues += 1
        if obligatoires_absentes:
            lot.rejets.append(RejetDemande(
                ligne=numero_ligne,
                motif="colonnes obligatoires absentes de l'export : "
                      + ", ".join(f"« {intitule} »" for intitule in obligatoires_absentes),
            ))
            continue
        valeurs: dict[str, object] = {}
        motif: str | None = None
        for index, question in colonnes.items():
            brut = ligne[index] if index < len(ligne) else None
            try:
                valeurs[question.champ] = _valeur_pour(question, brut)
            except ReponseIllisibleError as erreur:
                motif = str(erreur)
                break
        if motif is not None:
            lot.rejets.append(RejetDemande(ligne=numero_ligne, motif=motif))
            continue
        valeurs = {nom: valeur for nom, valeur in valeurs.items() if valeur is not None}
        try:
            lot.demandes.append(DemandeIAG.model_validate(valeurs))
        except ValidationError as erreur:
            # Les messages de pydantic citent le nom du champ, jamais la
            # réponse : le motif reste publiable dans un journal.
            champs = ", ".join(
                str(detail["loc"][0]) for detail in erreur.errors() if detail.get("loc")
            )
            lot.rejets.append(RejetDemande(
                ligne=numero_ligne,
                motif=f"réponse invalide ou manquante pour : {champs}",
            ))
    return lot
